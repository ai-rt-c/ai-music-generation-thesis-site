#!/usr/bin/env python3
"""Data pipeline for the thesis companion site.

Reads the dissertation spreadsheets and emits the JSON the website consumes.
Dev-time only: adds no runtime dependency to the Next.js app.

    python scripts/build_data.py

Outputs to ./data : papers.json systems.json evaluation.json audio-demos.json
                    taxonomy.json trends.json references.json content.json meta.json
"""
import json, os, re, sys
import openpyxl
from content_source import META, CONTENT, TAXONOMY, TRENDS

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, ".."))
OUT = os.path.join(ROOT, "data")
os.makedirs(OUT, exist_ok=True)


def private_source_path(env_name):
    value = os.environ.get(env_name)
    if not value:
        raise RuntimeError(
            f"{env_name} must point to a private local workbook outside the repository"
        )
    path = os.path.abspath(os.path.expanduser(value))
    try:
        inside_repository = os.path.commonpath([ROOT, path]) == ROOT
    except ValueError:
        inside_repository = False
    if inside_repository:
        raise RuntimeError(f"{env_name} must not point inside the public Git repository")
    if not os.path.isfile(path):
        raise FileNotFoundError(path)
    return path


MASTER = private_source_path("THESIS_MASTER_XLSX")
LISTEN = private_source_path("THESIS_LISTENING_XLSX")


def rows(path):
    wb = openpyxl.load_workbook(path)
    o = wb[wb.sheetnames[0]]
    hdr = [c.value for c in o[1]]
    out = []
    for r in o.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        out.append({hdr[i]: r[i] for i in range(len(hdr))})
    return out


def s(v):
    return "" if v is None else str(v).strip()


def num(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def norm_task(task):
    t = task.lower()
    if "arrang" in t or "accompan" in t:
        return "Arrangement"
    if "orchestr" in t:
        return "Orchestration"
    if "eval" in t or "benchmark" in t or "metric" in t:
        return "Evaluation"
    if "represent" in t:
        return "Representation"
    if "generation" in t or "composition" in t or "text-to-music" in t or "text-to-audio" in t:
        return "Generation"
    return "Other"


def norm_domain(rep):
    """Use the workbook's leading normalized representation label."""
    label = rep.strip().lower()
    if label.startswith("symbolic"):
        return "Symbolic"
    if label.startswith("audio"):
        return "Audio"
    if label.startswith("mixed"):
        return "Mixed"
    raise ValueError(f"unmapped data representation: {rep!r}")


def norm_primary_category(raw_category, domain):
    """Collapse the final corpus to the four approved analytical categories."""
    if raw_category in ("Arrangement", "Orchestration"):
        return raw_category
    if domain == "Symbolic":
        return "Symbolic generation"
    if domain == "Audio":
        return "Audio generation"
    raise ValueError(f"mixed-domain generative record needs manual classification")


def norm_paradigm(fam, method):
    f = fam.lower()
    m = (method or "").lower()
    tags = []
    if "diffusion" in f:
        tags.append("Diffusion")
    if "vae" in f:
        tags.append("VAE")
    if "gan" in f:
        tags.append("GAN")
    if "flow" in f:
        tags.append("Flow matching")
    if "rnn" in f or "lstm" in f:
        tags.append("RNN")
    if "state" in f or "mamba" in f or "state-space" in m or "mamba" in m:
        tags.append("State-space")
    if "transformer" in f or "llm" in f:
        tags.append("Transformer")
    if any(k in m for k in ["llama", "gpt", "large language model", "llm", "chatmusician", "pretrained language"]):
        if "LLM" not in tags:
            tags.append("LLM")
    if "foundation" in m or "foundation" in f:
        tags.append("Foundation model")
    if "hybrid" in f or "other" in f:
        if not tags:
            tags.append("Hybrid")
    if not tags:
        tags.append("Hybrid")
    # de-dup preserving order
    seen, out = set(), []
    for t in tags:
        if t not in seen:
            seen.add(t); out.append(t)
    return out[0], out


URL_RE = re.compile(r"https?://[^\s)]+")
REPO_RE = re.compile(r"(?:https?://)?((?:github\.com|gitlab\.com|huggingface\.co)/[^\s),;]+)")


def code_url(code):
    m = URL_RE.search(code or "")
    if m:
        return m.group(0).rstrip(".,);")
    m = REPO_RE.search(code or "")
    if m:
        return "https://" + m.group(1).rstrip(".,);")
    return None


def first_url(*vals):
    for v in vals:
        if not v:
            continue
        m = URL_RE.search(str(v))
        if m:
            return m.group(0).rstrip(".,;")
    return None


# ---------------- papers.json ----------------
master = rows(MASTER)
papers = []
for r in master:
    paper_id = str(r.get("ID"))
    task = s(r.get("Task"))
    rep = s(r.get("Data Representation"))
    fam = s(r.get("Architecture Family"))
    method = s(r.get("Method"))
    paradigm, ptags = norm_paradigm(fam, method)
    # MusicAgent is an LLM-powered planning/tool-selection agent, not flow matching.
    if paper_id == "360":
        paradigm, ptags = "Transformer", ["Transformer", "LLM"]
    domain = norm_domain(rep)
    task_category = norm_primary_category(norm_task(task), domain)
    code = s(r.get("Code"))
    demo = s(r.get("Availability of Demo"))
    papers.append({
        "id": paper_id,
        "title": s(r.get("Title")),
        "authors": s(r.get("Authors")),
        "year": int(r["Year"]) if r.get("Year") else None,
        "source": s(r.get("Source")),
        "task": task,
        "taskCategory": task_category,
        "domain": domain,
        "method": method,
        "architectureFamily": fam,
        "paradigm": paradigm,
        "paradigmTags": ptags,
        "dataRepresentation": rep,
        "dataset": s(r.get("Dataset")),
        "evaluation": s(r.get("Evaluation Method")),
        "metrics": s(r.get("Metrics")),
        "musicLength": s(r.get("Music Length")),
        "code": code,
        "hasCode": code.lower().startswith("yes"),
        "codeUrl": code_url(code),
        "hasDemo": demo.lower().startswith("yes"),
        "doi": s(r.get("DOI")) or None,
        "paperUrl": s(r.get("Paper URL")) or None,
        "inDepth": s(r.get("In-Depth Subset")).lower().startswith("yes"),
        "notes": s(r.get("Notes")),
        "citation": "",
    })

# ---------------- systems / evaluation / audio-demos (29) ----------------
listen = rows(LISTEN)
DIMS = {
    "Audio / Rendering Quality": "quality", "Melodic Coherence": "melody",
    "Harmonic Coherence": "harmony", "Rhythmic Stability": "rhythm",
    "Long-term Structure": "structure", "Condition / Control Adherence": "control",
    "Naturalness / Musicality": "naturalness", "Overall (1-5)": "overall",
}
STRENGTH = {
    "quality": "clean, high-fidelity rendering", "melody": "coherent, tuneful melodies",
    "harmony": "solid harmonic writing", "rhythm": "a stable rhythmic pulse",
    "structure": "convincing long-term structure", "control": "verifiable, faithful control",
    "naturalness": "natural, musical output",
}
WEAK = {
    "quality": "limited audio quality", "melody": "plain melodic writing",
    "harmony": "loose harmony", "rhythm": "unsteady rhythm",
    "structure": "weak long-term structure", "control": "control not clearly verifiable",
    "naturalness": "weak naturalness or vocals",
}
SHORT = {
    "quality": "audio quality", "melody": "melodic coherence", "harmony": "harmony",
    "rhythm": "rhythm", "structure": "long-term structure", "control": "control adherence",
    "naturalness": "naturalness",
}
paper_by_id = {p["id"]: p for p in papers}


def title_to_name(title):
    return title.split(":")[0].split("—")[0].strip()


def batch_index(batch):
    m = re.search(r"B(\d)", batch or "")
    return int(m.group(1)) if m else 0


def derive_strengths(scores):
    dims = [(k, v) for k, v in scores.items() if k != "overall" and isinstance(v, (int, float))]
    strong = sorted([d for d in dims if d[1] >= 4], key=lambda x: -x[1])
    if strong:
        return [STRENGTH[k].capitalize() if i == 0 else STRENGTH[k] for i, (k, _) in enumerate(strong[:3])]
    top = sorted(dims, key=lambda x: -x[1])[:2]
    return [("Relatively strong " + SHORT[k]) for k, _ in top]


def derive_weaknesses(scores):
    dims = [(k, v) for k, v in scores.items() if k != "overall" and isinstance(v, (int, float))]
    weak = sorted([d for d in dims if d[1] <= 3], key=lambda x: x[1])
    if weak:
        return [WEAK[k].capitalize() if i == 0 else WEAK[k] for i, (k, _) in enumerate(weak[:2])]
    return ["Few notable weaknesses within the evaluated dimensions"]


def derive_use_case(task, domain, scores):
    if task == "Arrangement":
        return "Generating accompaniment or arrangements from a given lead."
    if task == "Orchestration":
        return "Orchestration and instrumentation of existing material."
    ctrl = scores.get("control") or 0
    struct = scores.get("structure") or 0
    if ctrl >= 4:
        return "Controllable, prompt- or condition-driven generation."
    if struct >= 4:
        return "Longer-form, structurally coherent composition."
    if domain == "Audio":
        return "Text-to-music audio generation and prototyping."
    return "Symbolic music generation and prototyping."


systems, evaluation, demos = [], [], []
for r in listen:
    sid = str(r.get("ID"))
    title = s(r.get("Title"))
    scores = {}
    for col, key in DIMS.items():
        scores[key] = num(r.get(col))
    notes_blob = (s(r.get("Listening Notes")) + " " + s(r.get("Website Pick (1 sample)"))).lower()
    paper_based = "paper-reported" in notes_blob or "no public audio demo" in notes_blob
    p = paper_by_id.get(sid, {})
    systems.append({
        "id": sid, "paperId": sid, "name": title_to_name(title), "title": title,
        "year": int(r["Year"]) if r.get("Year") else p.get("year"),
        "batch": s(r.get("Batch")), "batchIndex": batch_index(s(r.get("Batch"))),
        "taskCategory": p.get("taskCategory", norm_task(s(r.get("Task")))),
        "domain": p.get("domain"), "paradigm": p.get("paradigm"),
    })
    notes = s(r.get("Listening Notes"))
    # split notes into technical contribution vs critical listening if a marker exists; else keep whole
    evaluation.append({
        "id": sid, "scores": scores,
        "technicalContribution": "", "criticalListening": notes, "takeaway": "",
        "paperBased": paper_based,
        "strengths": derive_strengths(scores),
        "weaknesses": derive_weaknesses(scores),
        "bestUseCase": derive_use_case(systems[-1]["taskCategory"], p.get("domain"), scores),
        "reportedMetrics": s(r.get("Reported Metrics (paper)")),
        "reportedResult": s(r.get("Reported Result (headline)")),
    })
    pick = s(r.get("Website Pick (1 sample)"))
    url = first_url(pick, r.get("Demo Link"))
    if paper_based or "no public" in pick.lower():
        dtype = "paper-only"
    elif "interactive" in pick.lower() or "hf space" in pick.lower() or "huggingface.co/spaces" in (url or ""):
        dtype = "interactive"
    elif url:
        dtype = "link"
    else:
        dtype = "paper-only"
    demos.append({
        "id": sid,
        "label": pick.split("—")[0].strip() if "—" in pick else (pick[:80] or "Demonstration"),
        "url": url, "type": dtype,
        "note": "Scores are paper-reported (no public demo)." if paper_based else "Hosted by the authors; opens in a new tab.",
    })

# ---------------- references.json (107 + 8 background) ----------------
def initials(g):
    return " ".join(t[0].upper() + "." for t in re.split(r"[\s\.]+", g.strip()) if t)


def one_author(pp):
    pp = pp.strip().strip(",").strip()
    if not pp:
        return None
    if "," in pp:
        sur, giv = pp.split(",", 1)
    else:
        tk = pp.split()
        if len(tk) == 1:
            return tk[0]
        sur, giv = tk[0], " ".join(tk[1:])
    ii = initials(giv.strip())
    return f"{sur.strip()}, {ii}" if ii else sur.strip()


def fmt_authors(a):
    a = str(a)
    etal = bool(re.search(r"(?i)\bet\.?\s*al\.?", a))
    a = re.sub(r"(?i)\bet\.?\s*al\.?", "", a)
    names = [one_author(x) for x in a.split(";")]
    names = [n for n in names if n]
    if not names:
        return "Unknown"
    if etal:
        return ", ".join(names) + ", et al."
    if len(names) == 1:
        return names[0]
    if len(names) == 2:
        return f"{names[0]}, & {names[1]}"
    return ", ".join(names[:-1]) + ", & " + names[-1]


def sortkey(a):
    a = str(a).split(";")[0].strip()
    return (a.split(",")[0] if "," in a else (a.split()[0] if a.split() else "zzz")).lower()


refs = []
paper_cite = {}
for r in master:
    au = fmt_authors(r.get("Authors"))
    yr = r.get("Year") or "n.d."
    title = s(r.get("Title")).rstrip(".")
    src = s(r.get("Source"))
    doi = s(r.get("DOI"))
    url = s(r.get("Paper URL"))
    tail = (doi if doi.startswith("http") else f"https://doi.org/{doi}") if doi and doi not in ("None", "N/A") else (url if url and url not in ("None", "N/A") else "")
    text = f"{au} ({yr}). {title}."
    if src and src not in ("None", "N/A"):
        text += f" {src}."
    if tail:
        text += f" {tail}"
    refs.append({"key": sortkey(r.get("Authors")), "text": text, "included": True})
    paper_cite[str(r.get("ID"))] = text

# attach the APA citation to each paper (single source: same generator as references)
for p in papers:
    p["citation"] = paper_cite.get(p["id"], "")

BG = [
    "Briot, J. P., Hadjeres, G., & Pachet, F. D. (2020). Deep learning techniques for music generation. Springer.",
    "Dash, A., & Agres, K. (2024). AI-based affective music generation systems: A review of methods and challenges. ACM Computing Surveys, 56(11), 1–34.",
    "Herremans, D., Chuan, C. H., & Chew, E. (2017). A functional taxonomy of music generation systems. ACM Computing Surveys, 50(5), 1–41.",
    "Ji, S., Yang, X., & Luo, J. (2023). A survey on deep learning for symbolic music generation: Representations, algorithms, evaluations, and challenges. ACM Computing Surveys, 56(1), 1–39.",
    "Page, M. J., McKenzie, J. E., Bossuyt, P. M., Boutron, I., Hoffmann, T. C., Mulrow, C. D., et al. (2021). The PRISMA 2020 statement: An updated guideline for reporting systematic reviews. BMJ, 372.",
    "Wang, L., Zhao, Z., Liu, H., Pang, J., Qin, Y., & Wu, Q. (2024). A review of intelligent music generation systems. Neural Computing and Applications, 36(12), 6381–6401.",
    "Yang, L. C., & Lerch, A. (2020). On the evaluation of generative models in music. Neural Computing and Applications, 32(9), 4773–4784.",
    "Zhu, Y., Baca, J., Rekabdar, B., & Rawassizadeh, R. (2023). A survey of AI music generation tools and models. arXiv preprint arXiv:2308.12982.",
]
for t in BG:
    refs.append({"key": t.split(",")[0].lower(), "text": t, "included": False})
refs.sort(key=lambda x: x["text"].lower())

# ---------------- write ----------------
def write(name, obj):
    with open(os.path.join(OUT, name), "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


write("papers.json", papers)
write("systems.json", systems)
write("evaluation.json", evaluation)
write("audio-demos.json", demos)
write("references.json", refs)
write("taxonomy.json", TAXONOMY)
write("trends.json", TRENDS)
write("content.json", CONTENT)
write("meta.json", META)

# ---------------- validate ----------------
assert len(papers) == 107, f"expected 107 papers, got {len(papers)}"
assert len(systems) == 29 and len(evaluation) == 29 and len(demos) == 29, "expected 29 systems"
assert len([p for p in papers if p["inDepth"]]) >= 29, "in-depth flag mismatch"
top = [e for e in evaluation if (e["scores"]["overall"] or 0) >= 4]
print(f"papers={len(papers)} systems={len(systems)} evaluations={len(evaluation)} demos={len(demos)}")
print(f"references={len(refs)} (included={sum(1 for r in refs if r['included'])})")
print(f"top systems (overall>=4)={len(top)}  taxonomy dims={len(TAXONOMY)}  trends={len(TRENDS)}")
print("wrote:", ", ".join(sorted(os.listdir(OUT))))
